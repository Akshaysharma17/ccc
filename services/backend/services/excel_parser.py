from io import BytesIO
from datetime import datetime

from openpyxl import load_workbook


def _str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    return str(val).strip() or None


def _find_sheet(wb):
    for name in wb.sheetnames:
        if name.lower().replace(" ", "") == "cccstatusmaster":
            return wb[name]
    for name in wb.sheetnames:
        if "ccc" in name.lower() and "master" in name.lower():
            return wb[name]
    raise ValueError(
        f"Sheet 'CCC status Master' not found. Available sheets: {wb.sheetnames}"
    )


def parse_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(BytesIO(file_bytes), read_only=False, data_only=True)
    ws = _find_sheet(wb)

    records = []
    # Forward-fill values for merged cells
    prev_section = None
    prev_section_name = None
    prev_kks_code = None
    prev_kks_desc = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=14, values_only=False), start=14):
        cells = [cell.value for cell in row]

        # Pad row to at least 55 columns
        while len(cells) < 55:
            cells.append(None)

        # Column mapping based on actual Excel layout (row 13 headers)
        # A(0): empty
        # B(1): Priority type
        # C(2): Department
        # D(3): Section group
        # E(4): Sub-code
        # F(5): Section number (1.1, 1.2, etc.)
        # G(6): Section name (Desal, etc.) — NEW column
        # H(7): Serial number
        # I(8): Sl No
        # J(9): KKS Code
        # K(10): KKS Description
        # L(11): CCC Number (full)
        # M(12): Part CCC no.
        # N(13): Part/Full
        # O(14): CCC Number (active)
        # P(15): Description (CCC description)
        # Q(16): Physical Work Completion Status
        # R(17): Erection Code
        # S(18): KKS Code of Equipment or System
        # T(19): Description (equipment)
        # U(20): Target Month of Submission
        # V(21): Responsibility
        # W(22): CCC Part/Full
        # X(23): Format 1 applicable
        # Y(24): Format 1 cleared
        # Z(25): Format 2 applicable
        # AA(26): Format 2 cleared
        # AB(27): Format 3 applicable
        # AC(28): Format 3 cleared
        # AD(29): Format 4 applicable
        # AE(30): Format 4 cleared
        # AF(31): Format 5 applicable
        # AG(32): Format 5 cleared
        # AH(33): CCC Status
        # AI(34): HD & Erection reports Available
        # AJ(35): HD Available
        # AK(36): Erection reports Available
        # AL(37): Engineering HO
        # AM(38): Procurement HO
        # AN(39): Quality HO
        # AO(40): Civil Team
        # AP(41): Mech Team
        # AQ(42): Electrical Team
        # AR(43): C&I Team
        # AS(44): Target Date of Submission

        priority_type = cells[1]
        department = cells[2]
        section_number = cells[5]
        section_name = cells[6]
        serial_number = cells[7]
        kks_code = cells[9]
        kks_description = cells[10]
        ccc_number_full = cells[11]
        ccc_number_part = cells[12]
        part_or_full = cells[13]
        active_ccc_number = cells[14]
        ccc_description = cells[15]
        ccc_status_raw = cells[33]

        # Formats
        fmt1_applicable = cells[23]
        fmt1_cleared = cells[24]
        fmt2_applicable = cells[25]
        fmt2_cleared = cells[26]
        fmt3_applicable = cells[27]
        fmt3_cleared = cells[28]
        fmt4_applicable = cells[29]
        fmt4_cleared = cells[30]
        fmt5_applicable = cells[31]
        fmt5_cleared = cells[32]

        # Reports
        hd_report = cells[35]
        erection_report_1 = cells[36]

        # Team handover
        team_engineering = cells[37]
        team_procurement = cells[38]
        team_quality = cells[39]
        team_civil = cells[40]
        team_mechanical = cells[41]
        team_electrical = cells[42]
        team_c_and_i = cells[43]

        target_date = cells[44]
        returned_status = cells[45]  # AT - Returned (sub-status for submitted CCCs)

        # Forward-fill for merged cells
        if section_number is not None:
            prev_section = _str(section_number)
        if section_name is not None:
            prev_section_name = _str(section_name)
        if kks_code is not None:
            prev_kks_code = _str(kks_code)
        if kks_description is not None:
            prev_kks_desc = _str(kks_description)

        # Skip rows with no system/section number in column F
        if section_number is None and prev_section is None:
            continue
        if section_number is None:
            continue

        # Skip empty rows (both serial number and CCC number are empty)
        if serial_number is None and ccc_number_full is None:
            continue

        # Normalize status: only "Pending" when cell is blank
        status = _str(ccc_status_raw)
        if status:
            status = status.title()
        else:
            status = "Pending"

        # Normalize returned_status (sub-status for submitted CCCs from column AT)
        returned_status_str = _str(returned_status)
        if returned_status_str:
            returned_status_str = returned_status_str.title()

        record = {
            "priority_type": _str(priority_type),
            "department": _str(department),
            "section_number": prev_section,
            "section_name": prev_section_name,
            "serial_number": _str(serial_number),
            "kks_code": prev_kks_code,
            "kks_description": prev_kks_desc,
            "ccc_number_full": _str(ccc_number_full),
            "ccc_number_part": _str(ccc_number_part),
            "part_or_full": _str(part_or_full),
            "active_ccc_number": _str(active_ccc_number),
            "ccc_description": _str(ccc_description),
            "format_1": {"applicable": _str(fmt1_applicable), "cleared": _str(fmt1_cleared)},
            "format_2": {"applicable": _str(fmt2_applicable), "cleared": _str(fmt2_cleared)},
            "format_3": {"applicable": _str(fmt3_applicable), "cleared": _str(fmt3_cleared)},
            "format_4": {"applicable": _str(fmt4_applicable), "cleared": _str(fmt4_cleared)},
            "format_5": {"applicable": _str(fmt5_applicable), "cleared": _str(fmt5_cleared)},
            "ccc_status": status,
            "returned_status": returned_status_str,
            "hd_report": _str(hd_report),
            "erection_report_1": _str(erection_report_1),
            "erection_report_2": None,
            "team_handover": {
                "engineering": _str(team_engineering),
                "procurement": _str(team_procurement),
                "quality": _str(team_quality),
                "civil": _str(team_civil),
                "mechanical": _str(team_mechanical),
                "electrical": _str(team_electrical),
                "c_and_i": _str(team_c_and_i),
            },
            "target_date": _str(target_date),
            "excel_row": row_idx,
        }
        records.append(record)

    wb.close()
    return records
